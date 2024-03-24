from typing import List, Dict, Union

from fastapi import APIRouter, Security, security, Depends, Query,BackgroundTasks
from fastapi.security import HTTPAuthorizationCredentials
from sqlmodel import select, Session, col
from starlette.responses import JSONResponse,FileResponse
from models import EmployeeWithServiceBase, Institution, Servicec, Session, Employee,EmployeeBase,Contract,EmployeeS,Service
from fastapi import FastAPI, status, UploadFile,File
from database import engine
from auth import AuthHandler
import shutil
import os
import sqlalchemy
import json
import openpyxl
from base_models.invoices import FetchEmployees
from utils.excel import generate_excel
from base_models.employee import EmployeBase
from openpyxl.styles import Font,Alignment
emp_router = APIRouter()
# emp_session=Session(bind=engine)

auth_handler = AuthHandler()


def get_session():
    with Session(engine) as session:
        yield session


@emp_router.post(
    "/institutions/{institution_id}/employee/save",
    response_model=Employee,
    tags=["Employees"],
)
async def create_institution_employee(
    institution_id: int,
    emp: EmployeeBase,
    emp_session: Session = Depends(get_session),
    user=Depends(auth_handler.get_current_user),
):
    """Endpoint to Create Institution Employee"""

    try:
        statement = select(Institution).where(
            Institution.id == institution_id, Institution.deletedStatus == False
        )
        institution = emp_session.exec(statement).first()

        if not institution is None:
           # new_service_list=[]
                #new_service_list.append(context)
            new_emp = Employee(
                firstName=emp.firstName,
                lastName=emp.lastName,
                gender=emp.gender,
                email=emp.email,
                phone=emp.phone,
                title=emp.title,
                institution_id=institution_id,
                referenceName=institution.name,
                active_status="Active",
                services=emp.services
            )
            emp_session.add(new_emp)
            emp_session.commit()

            # file_path='EmployeeImages/'+new_emp.uuid+'_'+file.filename
            # with open(file_path,'wb') as buffer:
            #     shutil.copyfileobj(file.file,buffer)

            # new_emp.photo =file_path
            # new_emp.institution_id=institution_id
            # emp_session.add(new_emp)
            # emp_session.commit()
            return new_emp
        else:
            return JSONResponse(
                content="Institution with " + str(institution_id) + " Not Found",
                status_code=status.HTTP_404_NOT_FOUND,
            )
    except Exception as e:
        print(e)
        return JSONResponse(
            content="Error: " + str(e),
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@emp_router.get("/employees/{emp_id}/", response_model=Employee, tags=["Employees"])
async def fetch_employee_detail(
    emp_id: int,
    emp_session: Session = Depends(get_session),
    user=Depends(auth_handler.get_current_user),
):
    """Endpoint to Return Employee Detail"""

    try:
        statement = select(Employee).where(
            Employee.id == emp_id, Employee.deletedStatus == False
        )
        result = emp_session.exec(statement).first()
        if result is not None:
            return result
        else:
            return JSONResponse(
                content="Employee with " + str(emp_id) + " Not Found",
                status_code=status.HTTP_404_NOT_FOUND,
            )

    except Exception as e:
        print(e)
        return JSONResponse(
            content="Error: " + str(e),
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

""""
@emp_router.put(
    "/employees/{emp_id}/update", response_model=Employee, tags=["Employees"]
)
async def update_employee(
    emp_id: int,
    emp: EmployeeBase,
    emp_session: Session = Depends(get_session),
    user=Depends(auth_handler.get_current_user),
):
    #Endpoint to Update  Employee Data"""
"""""
    try:
        statement = select(Employee).where(
            Employee.id == emp_id, Employee.deletedStatus == False
        )
        result = emp_session.exec(statement).first()
        
    
        
        result.firstName = emp.firstName
        result.lastName = emp.lastName
        result.email = emp.email
        result.gender = emp.gender
        result.phone = emp.phone
        result.title = emp.title
        result.services=emp.services

        emp_session.add(result)
        emp_session.commit()
    return result
    else:
    return JSONResponse(
    content="Employee with " + str(emp_id) + " Not Found",
            status_code=status.HTTP_404_NOT_FOUND,
            )
except Exception as e:
     print(e)
    return JSONResponse(
            content="Error: " + str(e),
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )
"""


@emp_router.get("/institutions/{institution_id}/employees/", tags=["Employees"])
async def fetch_employees_of_institution(
    institution_id: int,
    emp_session: Session = Depends(get_session),
    user=Depends(auth_handler.get_current_user),
):

    """Endpoint to Fetch Institution Employees"""

    statement = select(Employee).where(
        Employee.institution_id == institution_id, Employee.deletedStatus == False
    )
    results = emp_session.exec(statement).all()

    return results


@emp_router.delete(
    "/employees/{emp_id}/delete/", response_model=Employee, tags=["Employees"]
)
async def delete_employee(
    emp_id: int,
    emp_session: Session = Depends(get_session),
    user=Depends(auth_handler.get_current_user),
):
    """Endpoint to delete a Employee"""

    try:
        statement = select(Employee).where(
            Employee.id == emp_id, Employee.deletedStatus == False
        )
        result = emp_session.exec(statement).first()

        if not result is None:
            result.deletedStatus = True
            emp_session.add(result)
            emp_session.commit()

            return result
        else:
            return JSONResponse(
                content="Employee with " + str(emp_id) + " Not Found",
                status_code=status.HTTP_404_NOT_FOUND,
            )
    except Exception as e:
        print(e)
        return JSONResponse(
            content="Error: " + str(e),
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@emp_router.put(
    "/employees/{emp_id}/uplod_photo/", response_model=Employee, tags=["Employees"]
)
async def upload_employee_image(
    emp_id: int,
    file: UploadFile,
    emp_session: Session = Depends(get_session),
    user=Depends(auth_handler.get_current_user),
):
    """Endpoint to Update  Employee Data"""

    try:
        statement = select(Employee).where(
            Employee.id == emp_id, Employee.deletedStatus == False
        )
        result = emp_session.exec(statement).first()

        if not result is None:
            if os.path.exists(os.environ["FILE_SOURCE"] + "/" + "EmployeeImages"):
                file_path = (
                    os.environ["FILE_SOURCE"]
                    + "/"
                    + "EmployeeImages/"
                    + str(result.uuid)
                    + "_"
                    + file.filename
                )
                with open(f"{file_path}", "wb") as buffer:
                    shutil.copyfileobj(file.file, buffer)
            else:
                os.mkdir(os.environ["FILE_SOURCE"] + "/" + "EmployeeImages")
                file_path = (
                    os.environ["FILE_SOURCE"]
                    + "/"
                    + "EmployeeImages/"
                    + str(result.uuid)
                    + "_"
                    + file.filename
                )
                with open(f"{file_path}", "wb") as buffer:
                    shutil.copyfileobj(file.file, buffer)
            result.photo = file_path
            emp_session.add(result)
            emp_session.commit()
            return result
        else:
            return JSONResponse(
                content="Employee with " + str(emp_id) + " Not Found",
                status_code=status.HTTP_404_NOT_FOUND,
            )
    except Exception as e:
        print(e)
        return JSONResponse(
            content="Error: " + str(e),
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@emp_router.get("/employees/all", tags=["Employees"])
async def fetch_all_employees(
    emp_session: Session = Depends(get_session),
    user=Depends(auth_handler.get_current_user),
):

    """Endpoint to Fetch All Employees"""

    statement = select(Employee).where(
        Employee.deletedStatus == False, Employee.active_status == "Active"
    )
    results = emp_session.exec(statement).all()

    return results


@emp_router.put(
    "/employees/{emp_id}/activate/", response_model=Employee, tags=["Employees"]
)
async def activate_employee(
    emp_id: int,
    emp_session: Session = Depends(get_session),
    user=Depends(auth_handler.get_current_user),
):
    """Endpoint to Activate a Employee"""

    try:
        statement = select(Employee).where(
            Employee.id == emp_id, Employee.deletedStatus == False
        )
        result = emp_session.exec(statement).first()

        if not result is None:
            result.active_status = "Active"
            emp_session.add(result)
            emp_session.commit()

            return result
        else:
            return JSONResponse(
                content="Employee with " + str(emp_id) + " Not Found",
                status_code=status.HTTP_404_NOT_FOUND,
            )
    except Exception as e:
        print(e)
        return JSONResponse(
            content="Error: " + str(e),
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@emp_router.put(
    "/employees/{emp_id}/deactivate/", response_model=Employee, tags=["Employees"]
)
async def deactivate_employee(
    emp_id: int,
    emp_session: Session = Depends(get_session),
    user=Depends(auth_handler.get_current_user),
):
    """Endpoint to Deactivate a Employee"""

    try:
        statement = select(Employee).where(
            Employee.id == emp_id, Employee.deletedStatus == False
        )
        result = emp_session.exec(statement).first()

        if not result is None:
            result.active_status = "Inactive"
            emp_session.add(result)
            emp_session.commit()

            return result
        else:
            return JSONResponse(
                content="Employee with " + str(emp_id) + " Not Found",
                status_code=status.HTTP_404_NOT_FOUND,
            )
    except Exception as e:
        print(e)
        return JSONResponse(
            content="Error: " + str(e),
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@emp_router.get("/employees/{search_param}/details", tags=["Employees"])
async def search_employee_detail(
    search_param: str,
    emp_session: Session = Depends(get_session),
    user=Depends(auth_handler.get_current_user),
):
    """Endpoint to Return Employee Detail"""

    try:
        emp_list = []
        full_names = Employee.firstName + " " + Employee.lastName
        statement = select(Employee).where(
            sqlalchemy.func.lower(full_names).contains(search_param.lower()),
            # .where(func.lower(col(Transaction.buyerName)).contains(name.lower()))
            Employee.deletedStatus == False,
        )
        print('-------')
        result = emp_session.exec(statement).all()
        if len(result) > 0:
            for r in result:
                inst_statement = select(Institution).where(
                    Institution.id == r.institution_id, Employee.deletedStatus == False
                )
                ins = emp_session.exec(inst_statement).first()
                services=[]
                if not r.services is None:
                    services=r.services
                data = {
                    "uuid": str(r.uuid),
                    "firstName": r.firstName,
                    "lastName": r.lastName,
                    "gender": r.gender,
                    "email": r.email,
                    "phone": r.phone,
                    "active_status": r.active_status,
                    "institutionName": ins.name,
                    "institutionEmail": ins.email,
                    "institutionPhone": ins.phone,
                    "services":services
                }
                emp_list.append(data)
            return emp_list

        else:
            return JSONResponse(
                content="Employee search Not Found",
                status_code=status.HTTP_404_NOT_FOUND,
            )

    except Exception as e:
        print(e)
        return JSONResponse(
            content="Error: " + str(e),
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    
@emp_router.put(
    "/employee/{employee_id}/update-services",
    response_model=Employee,
    tags=["Employees"],
)
async def update_employee_services(
    employee_id:int, emp:EmployeeS,
    emp_session: Session = Depends(get_session),
    user=Depends(auth_handler.get_current_user),
):
    #Endpoint to update Employee services

    try:
        statement = select(Employee).where(
            Employee.id==employee_id, Employee.deletedStatus == False
        )
        employee = emp_session.exec(statement).first()
        if  not employee is None:
            inst_statement=select(Institution).where(Institution.id==employee.institution_id)
            institution=emp_session.exec(inst_statement).first()
            if not institution  is None:
                cont_statement=select(Contract).where(Contract.institution_id==institution.id,
                                                     Contract.current==True)
            contract=emp_session.exec(cont_statement).first()
            new_serv_list=[]
            for s in emp.services:
                check=False
                for servc in contract.services:
                    if servc['service_id']==s.service_id:
                        check=True
                        break
                if check:
                    servi_i=s.service_id
                    servi_r=s.service_rate
                    #servi_name=s.service_name
                    statements=select(Service).where(Service.id==servi_i)
                    service=emp_session.exec(statements).first()
                    servi_name=service.name
                    context={
                    "service_id":servi_i,
                    "service_rate":servi_r,
                    "name":servi_name

                   }
                    new_serv_list.append(context)
            
            employee.services=new_serv_list
            emp_session.add(employee)
            emp_session.commit()

            return employee
        else:
            return JSONResponse(
                content="EMployee with " + str(employee_id) + " Not Found",
                status_code=status.HTTP_404_NOT_FOUND,
            )
    except Exception as e:
        print(e)
        return JSONResponse(
            content="Error: " + str(e),
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )





# Method to find employee


def getEmployee(emp_id):
    with Session(engine) as emp_session:
        st = select(Employee).where(
            Employee.id == emp_id, Employee.deletedStatus == False
        )
        emp = emp_session.exec(st).first()
        return emp

@emp_router.post('/institutions/{institution_id}/employee/save_with_service',response_model=Employee,tags=["Employees"])
async def create_institution_employee_with_service(institution_id:int,emp:EmployeeWithServiceBase,emp_session: Session = Depends(get_session),user=Depends(auth_handler.get_current_user)):
    """ Endpoint to Create Institution Employee with service """ 
    try:
        statement = select(Institution).where(Institution.id==institution_id,Institution.deletedStatus==False)
        institution = emp_session.exec(statement).first()
        if not institution is None:
            statement1 = select(Contract).where(Contract.institution_id==institution_id,Contract.deletedStatus == False, Contract.current==True,
             )
            contract = emp_session.exec(statement1).first()
            service_list=[]
            for s in emp.services:
                # Find service in contract service
                check=False
                for cs in contract.services:
                    if cs['service_id']==s.service_id:
                        check=True
                        break
                if check:
                    service_i=s.service_id
                    service_r=s.service_rate
                    states=select(Service).where(Service.id==service_i)
                    serv=emp_session.exec(states).first()
                    service_n=serv.name
                    context={
                        "service_id":service_i,
                        "service_rate":service_r,
                        "name":service_n
                    }
                    service_list.append(context)
            new_employee =Employee(firstName=emp.firstName,lastName=emp.lastName,email=emp.email,gender=emp.gender,phone=emp.phone,title=emp.title,institution_id=institution_id,referenceName=institution.name,services=service_list,active_status='Active')
            emp_session.add(new_employee)
            emp_session.commit()

            return new_employee
        else:
            return JSONResponse(content="Institution with "+institution_id+" Not Found",status_code=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        print(e)
        return JSONResponse(content="Error: "+str(e),status_code=status.HTTP_500_INTERNAL_SERVER_ERROR)




@emp_router.get("/employees/{employee_id}/assigned_and_unassigned_services",response_model=List[Servicec], tags=["Employees"])
async def get_employee_assigned_and_unassigned_services(
    employee_id: int,
    emp_session: Session = Depends(get_session),
    user=Depends(auth_handler.get_current_user),
):
    """Endpoint to Return Employee assigned  and unassigned services"""

    try:
       
        
        statement = select(Employee).where(Employee.id==employee_id,Employee.deletedStatus == False,
        )
        emp = emp_session.exec(statement).first()
        
        if not emp is None:
            list=[]
            statement1 = select(Contract).where(Contract.institution_id==emp.institution_id,Contract.deletedStatus == False, Contract.current==True,
             )
            contract = emp_session.exec(statement1).first()
            if not contract.services is None:
                for cs in contract.services:
                    service=Servicec()
                    service.service_id=cs['service_id']
                    service.service_name=cs['name']
                    service.service_rate=cs['service_rate']
                    check=False
                    if not emp.services is None:
                        for s in emp.services:
                            if cs['service_id']==s['service_id']:
                                check=True
                                break
                    
                    if check:
                        service.assigned=True
                    else:
                        service.assigned=False

                    list.append(service)
                
            return list
            
        else:
            return JSONResponse(content="Employee with "+str(employee_id)+" Not Found",status_code=status.HTTP_404_NOT_FOUND)

    except Exception as e:
        print(e)
        return JSONResponse(
            content="Error: " + str(e),
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )
    

@emp_router.get("/employees/{emp_id}/services", tags=["Employees"])
async def fetch_employee_services(
    emp_id: int,
    emp_session: Session = Depends(get_session),
    user=Depends(auth_handler.get_current_user),
):
    """Endpoint to Return Employee services"""

    try:
        statement = select(Employee).where(Employee.id == emp_id, Employee.deletedStatus == False)
        result = emp_session.exec(statement).first()
        if result is not None:
            serv=result.services

            return serv
        else:
            return JSONResponse(
                content="Employee with " + str(emp_id) + " Not Found",
                status_code=status.HTTP_404_NOT_FOUND,
            )

    except Exception as e:
        print(e)
        return JSONResponse(
            content="Error: " + str(e),
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )
    
    

# Save Business documents
@emp_router.post(
    "/employees/institution/{institution_uuid}/uploadEmployeer",  tags=["Employees"]
)
async def save_supporting_document(
    institution_uuid:str,
    file: UploadFile = File(...),
    emp_session: Session = Depends(get_session),
   user=Depends(auth_handler.get_current_user),
):
    """Endpoint to Create Employee from excel file with institution_uuid"""
    try:
       statement = select(Institution).where(Institution.uuid == str(institution_uuid), Institution.deletedStatus == False)
       result = emp_session.exec(statement).first()
       if result is None:
            return JSONResponse(
                content="Institution with " + str(institution_uuid) + " Not Found",
                status_code=status.HTTP_404_NOT_FOUND,
            )
       else:
            if os.path.exists(os.environ["FILE_SOURCE"] + "/" + "EmployeeUploadedFile/"):
                file_path = (
                    os.environ["FILE_SOURCE"]
                    + "/"
                    + "EmployeeUploadedFile/"
                    + str(institution_uuid)
                    + "_"
                    + file.filename
                )
                with open(f"{file_path}", "wb") as buffer:
                    shutil.copyfileobj(file.file, buffer)
            else:
                os.mkdir(os.environ["FILE_SOURCE"] + "/" + "EmployeeUploadedFile")
                file_path = (
                    os.environ["FILE_SOURCE"]
                    + "/"
                    + "EmployeeUploadedFile/"
                    + str(institution_uuid)
                    + "_"
                    + file.filename
                )
                with open(f"{file_path}", "wb") as buffer:
                    shutil.copyfileobj(file.file, buffer)
            print(file_path)
            wb = openpyxl.load_workbook(file_path)
            sheet_obj = wb.active
            list=[]
            for row in sheet_obj.iter_rows(min_row=2):
                employee=Employee()
                if not row[0].value is None:
                    employee.firstName=row[0].value
                if not row[1].value is None:
                    employee.lastName=row[1].value
                if not row[2].value is None:
                    employee.gender=row[2].value
                if not row[3].value is None:
                    employee.email=row[3].value
                if not row[4].value is None:
                    employee.phone=row[4].value
                if not row[5].value is None:
                    employee.title=row[5].value

                #Save employee in db
                employee.institution_id=result.id
                employee.referenceName=result.name
                employee.active_status="Active"
                emp_session.add(employee)
                emp_session.flush()
                list.append(employee)
            emp_session.commit()
            #remove file 
            if os.path.exists(file_path):
                os.remove(file_path)
            return JSONResponse(
                content=""+str(len(list))+" Employee(s) Successfully saved",
                status_code=status.HTTP_201_CREATED,
            )
    except Exception as e:
        print(e)
        return JSONResponse(
            content="Error: " + str(e),
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

@emp_router.post(
    "/employees/institution/{institution_id}/uploadEmployees",  tags=["Employees"]
)
async def save_employee_file_document(
    institution_id:int,
    file: UploadFile = File(...),
    emp_session: Session = Depends(get_session),
    user=Depends(auth_handler.get_current_user),
):
    """Endpoint to Create Employee from excel file with institution_id"""
    try:
       statement = select(Institution).where(Institution.id == institution_id, Institution.deletedStatus == False)
       result = emp_session.exec(statement).first()
       if result is None:
            return JSONResponse(
                content="Institution with " + str(institution_id) + " Not Found",
                status_code=status.HTTP_404_NOT_FOUND,
            )
       else:
            if os.path.exists(os.environ["FILE_SOURCE"] + "/" + "EmployeeUploadedFile/"):
                file_path = (
                    os.environ["FILE_SOURCE"]
                    + "/"
                    + "EmployeeUploadedFile/"
                    + str(result.uuid)
                    + "_"
                    + file.filename
                )
                with open(f"{file_path}", "wb") as buffer:
                    shutil.copyfileobj(file.file, buffer)
            else:
                os.mkdir(os.environ["FILE_SOURCE"] + "/" + "EmployeeUploadedFile")
                file_path = (
                    os.environ["FILE_SOURCE"]
                    + "/"
                    + "EmployeeUploadedFile/"
                    + str(result.uuid)
                    + "_"
                    + file.filename
                )
                with open(f"{file_path}", "wb") as buffer:
                    shutil.copyfileobj(file.file, buffer)
            print(file_path)
            wb = openpyxl.load_workbook(file_path)
            sheet_obj = wb.active
            list=[]
            for row in sheet_obj.iter_rows(min_row=2):
                employee=Employee()
                if not row[0].value is None:
                    employee.firstName=row[0].value
                if not row[1].value is None:
                    employee.lastName=row[1].value
                if not row[2].value is None:
                    employee.gender=row[2].value
                if not row[3].value is None:
                    employee.email=row[3].value
                if not row[4].value is None:
                    employee.phone=row[4].value
                if not row[5].value is None:
                    employee.title=row[5].value

                #Save employee in db
                employee.institution_id=result.id
                employee.referenceName=result.name
                employee.active_status="Active"
                emp_session.add(employee)
                emp_session.flush()
                list.append(employee)
            emp_session.commit()
            #remove file 
            if os.path.exists(file_path):
                os.remove(file_path)
            return JSONResponse(
                content=""+str(len(list))+" Employee(s) Successfully saved",
                status_code=status.HTTP_201_CREATED,
            )
    except Exception as e:
        print(e)
        return JSONResponse(
            content="Error: " + str(e),
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

@emp_router.get("/institutions/{institution_id}/download_excel/employees/", tags=["Employees"])
async def download_employees_of_institution(
    institution_id: int,fetch:FetchEmployees,bg_tasks: BackgroundTasks,
    emp_session: Session = Depends(get_session),
    user=Depends(auth_handler.get_current_user),
):

    """Endpoint to download Institution Employees on excel"""

    statement = select(Employee).where(
        Employee.institution_id == institution_id, Employee.deletedStatus == False
    )
    excel_list=[]
    results = emp_session.exec(statement).all()
    if len(results)>0:  
        for r in results:
          excel_list.append(EmployeBase(firstName=r.firstName,
                                       lastName=r.lastName,
                                       gender=r.gender,
                                       email=r.email,
                                       phone=r.phone,
                                       title=r.title,
                                       ))
                                    
        #print('11111111111111111111111111111',excel_list)
    excel_list.sort(key=lambda x: x.firstName)
    if fetch == FetchEmployees.download:
            headers = ['firstName', 'lastName', 'gender', 'email', 'phone', 'title']
            file_name = f"{r.referenceName}"
            generate_excel(excel_list, headers, file_name)
            headers = {'Content-Disposition': f'attachment; filename="{file_name}.xlsx"'}
            file_path = os.getcwd() + "/" + f'{file_name}.xlsx'

            bg_tasks.add_task(os.remove, file_path)
            return FileResponse(path=file_path, media_type='application/octet-stream',
                                filename=f'{file_name}.xlsx', headers=headers, background=bg_tasks)
    return excel_list
    


@emp_router.get("/institutions/{institution_id}/download/employees/", tags=["Employees"])
async def download_employees_of_institution_file(
    institution_id: int,fetch:FetchEmployees,bg_tasks: BackgroundTasks,
    emp_session: Session = Depends(get_session),
    user=Depends(auth_handler.get_current_user),
):

    """Endpoint to download Institution Employees on excel"""

    statement = select(Employee).where(
        Employee.institution_id == institution_id, Employee.deletedStatus == False
    )

    results = emp_session.exec(statement).all()
    if len(results)>0:
        wb = openpyxl.Workbook()
        ws = wb.active

        headers=['FirstName', 'LastName', 'Gender', 'Email', 'Phone', 'Title']
        ws.append(headers)
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = header_alignment

        for r in results:
            employe = (r.firstName, r.lastName, r.gender, r.email, r.phone, r.title)
            ws.append(employe)
            
        ws.insert_rows(1)    
        ws.merge_cells('A1:F1')
        ws['A1'] = 'List of '+str(r.referenceName)+' Employees'
        ws['A1'].alignment = header_alignment

        file_name = 'employee.xlsx'
        file_path = os.getcwd() + "/" + file_name
        wb.save(file_path)

        if fetch == FetchEmployees.download:
            bg_tasks.add_task(os.remove, file_path)
            return FileResponse(path=file_path, media_type='application/octet-stream',
                                filename=file_name, background=bg_tasks)
    return wb
  
