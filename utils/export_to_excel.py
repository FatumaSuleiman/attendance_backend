from sqlmodel import select, Session
from database import engine
from models import Employee
import openpyxl
from openpyxl.styles import Font, Alignment
from typing import List


def export_to_excel(entries_input: List, filename: str):
    """ FUNCTION TO EXPORT DATA TO EXCEL """

    # declare bold font and center alignment style
    bold_font = Font(bold=True, size=12)
    center_aligned_text = Alignment(horizontal="center")

    # Initiate Excel sheet
    book = openpyxl.Workbook()
    sheet = book.active
    dest_filename = f'{filename}.xlsx'
    headers_row = ['names', 'institution', 'branch', 'attendance date']
    sheet.append(headers_row)

    for row in sheet.iter_cols():
        for cell in row:
            col_letter = str(cell).strip("(<Cell 'Sheet'").strip('.1>')
            cell.value = cell.value.title()
            cell.font = bold_font
            cell.alignment = center_aligned_text
            sheet.column_dimensions[f"{col_letter}"].width = 20

    i = 1
    for row in entries_input:
        row = row.dict()
        i += 1
        j = 1

        dict_keys = ('names', 'institution', 'branch', 'attendance_date')
        new_dict = dict.fromkeys(dict_keys)

        # Get employee data
        with Session(engine) as session:
            statement = select(Employee).where(Employee.id == row['employee_id'],
                                               Employee.deletedStatus == False)
            employee_data = session.exec(statement).first()
            new_dict['names'] = f"{employee_data.firstName} {employee_data.lastName}"
            new_dict['institution'] = employee_data.referenceName
            new_dict['branch'] = row['branchName']
            new_dict['attendance_date'] = row['signedAt']

        for col in new_dict:
            try:
                cell = sheet.cell(row=i, column=j)
                cell.value = new_dict[f'{col}']
                j += 1
            except Exception as e:
                return str(e)

    # SAVE EXCEL FILE
    book.save(filename=dest_filename)
    return book
