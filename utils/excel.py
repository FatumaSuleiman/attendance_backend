import openpyxl
from openpyxl.styles import Font, Alignment
from typing import List


def generate_excel(entries_input: List, headers_row: List, filename: str):
    """ FUNCTION TO EXPORT DATA TO EXCEL """

    # declare bold font and center alignment style
    bold_font = Font(bold=True, size=12)
    center_aligned_text = Alignment(horizontal="center")

    # Initiate Excel sheet
    book = openpyxl.Workbook()
    sheet = book.active
    dest_filename = f'{filename}.xlsx'
    sheet.append(headers_row)

    for row in sheet.iter_cols():
        for cell in row:
            col_letter = str(cell).strip("(<Cell 'Sheet'").strip('.1>')
            cell.value = cell.value.title()
            cell.font = bold_font
            cell.alignment = center_aligned_text
            sheet.column_dimensions[f"{col_letter}"].width = 20

    i = 1
    for row in entries_input:
        data_list = list(row.dict().values())
        i += 1
        j = 1
        dict_keys = tuple(headers_row)
        new_dict = dict.fromkeys(dict_keys)

        for d, column in enumerate(dict_keys):
            new_dict[f'{column}'] = data_list[d]
        for col in new_dict:
            try:
                cell = sheet.cell(row=i, column=j)
                cell.value = new_dict[f'{col}']
                j += 1
            except Exception as e:

                return str(e)

    # SAVE EXCEL FILE
    book.save(filename=dest_filename)
    return book
